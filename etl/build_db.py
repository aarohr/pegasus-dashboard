"""
build_db.py
-----------
One-time ETL: migrate the Project Pegasus QofE databook (Excel) into a SQLite
database. The Streamlit app reads ONLY from the database, never from Excel.

Run:  python etl/build_db.py
Output: data/pegasus.db

Row/column coordinates below were mapped against the blinded databook
(2_4_1_Project_Pegasus ... .xlsx). The workbook layout is fixed, so the
coordinates are hard-coded with reconciliation checks that fail loudly if a
future version of the workbook shifts.
"""
from __future__ import annotations
import os
import sqlite3
import warnings
import datetime as dt
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "pegasus_databook.xlsx"   # source workbook
DB = ROOT / "data" / "pegasus.db"

# ---- location taxonomy (Clinic vs ASC, TX vs GA) -------------------------
ASC_SITES = {"Tyler Surgery Center", "ASC-TEXARKANA"}
TX_SITES = {
    "PSC-TYLER", "ASC-TEXARKANA", "PSC-TEXARKANA", "Tyler Surgery Center",
    "PSC-LONGVIEW", "PSC-LUFKIN", "[MD 1] PC",
}
GA_SITES = {
    "NSO-MARIETTA", "NSO-CONYERS", "CSP-MARIETTA", "CSP-CONYERS",
    "MDP-CONYERS", "NSO-CARROLLTON",
}
SKIP_ROWS = {"Texas", "Georgia", "Total", "CORPORATE", "Eliminations",
             "[Management Co.]", "Georgia - DCO", "Roseland", "Excalibur"}


def _months(n_years_start=2023):
    """12 monthly periods per year, 2023-2025, as YYYY-MM strings by index."""
    return None  # helper not needed; we read dates from the sheet


def load():
    if not SRC.exists():
        raise FileNotFoundError(
            f"Source workbook not found at {SRC}. "
            "Place the QofE databook there (see README)."
        )
    return load_workbook(SRC, read_only=True, data_only=True)


# --------------------------------------------------------------------------
def kpi_annual(wb) -> pd.DataFrame:
    """5 KPIs + core revenue metrics from DD1. Quality of Revenue (annual)."""
    ws = wb["DD1. Quality of Revenue"]
    rows = list(ws.iter_rows(values_only=True))
    C = {2023: 55, 2024: 56, 2025: 57}          # annual columns
    R = {
        "reported_net_revenue": 50,
        "net_revenue": 70,                       # Net Revenue, Adjusted (QofE basis)
        "cash_collections": 65,                  # Date of Service Cash Collections
        "est_future_collections": 66,
        "gross_charges": 77,
        "procedure_count": 78,
        "encounters": 79,
        "rev_per_procedure": 87,                 # Adjusted revenue per procedure
    }
    recs = []
    for yr, col in C.items():
        rec = {"year": yr}
        for k, ri in R.items():
            rec[k] = rows[ri][col]
        rec["gross_to_net_pct"] = rec["net_revenue"] / rec["gross_charges"] * 100
        rec["cash_collection_rate_pct"] = rec["cash_collections"] / rec["gross_charges"] * 100
        rec["efc_pct_of_net_rev"] = rec["est_future_collections"] / rec["net_revenue"] * 100
        recs.append(rec)
    df = pd.DataFrame(recs)
    # reconciliation
    r25 = df[df.year == 2025].iloc[0]
    assert abs(r25.net_revenue - 35252.6) < 5, "DD1 net revenue tie broke"
    assert abs(r25.gross_to_net_pct - 31.7) < 0.3, "gross-to-net tie broke"
    return df


def pnl_monthly(wb) -> pd.DataFrame:
    """Monthly revenue + gross profit (consolidated IS) + EBITDA (QofE)."""
    # revenue + gross profit from Clinic v. Corporate -> Consolidated block
    cvc = list(wb["Clinic v. Corporate"].iter_rows(values_only=True))
    hdr = cvc[8]
    cons_cols = list(range(90, 126))             # consolidated monthly Jan23-Dec25
    months = [hdr[j] for j in cons_cols]
    net_rev = [cvc[54][j] for j in cons_cols]
    gp = [cvc[85][j] for j in cons_cols]

    # EBITDA from QofE Table monthly cols 3-38
    q = list(wb["QofE Table"].iter_rows(values_only=True))
    qcols = list(range(3, 39))
    rep_ebitda = [q[32][j] for j in qcols]
    dd_adj = [q[47][j] for j in qcols]
    interest = [q[26][j] for j in qcols]

    df = pd.DataFrame({
        "month": [m.strftime("%Y-%m") if isinstance(m, dt.datetime) else m for m in months],
        "net_revenue": net_rev,
        "gross_profit": gp,
        "reported_ebitda": rep_ebitda,
    })
    df["adjusted_ebitda"] = [a + b for a, b in zip(rep_ebitda, dd_adj)]
    df["interest_expense"] = interest
    df["gross_margin_pct"] = df.gross_profit / df.net_revenue * 100
    df["ebitda_margin_pct"] = df.adjusted_ebitda / df.net_revenue * 100
    df["date"] = pd.to_datetime(df.month + "-01")
    return df.sort_values("date").reset_index(drop=True)


def location_annual(wb) -> pd.DataFrame:
    """Net revenue by location/year from Location Summary (Adjusted block)."""
    ws = list(wb["Location Summary"].iter_rows(values_only=True))
    # Adjusted block header at row 50; data rows 51-94. months cols 2-37.
    yr_cols = {2023: range(2, 14), 2024: range(14, 26), 2025: range(26, 38)}
    recs = []
    for i in range(51, 95):
        if i >= len(ws):
            break
        lbl = ws[i][1]
        if not isinstance(lbl, str) or not lbl.strip():
            continue
        lbl = lbl.strip()
        if lbl in SKIP_ROWS:
            continue
        if lbl not in (TX_SITES | GA_SITES):
            continue
        for yr, cols in yr_cols.items():
            val = sum(ws[i][j] for j in cols if isinstance(ws[i][j], (int, float)))
            recs.append({
                "location": lbl,
                "state": "TX" if lbl in TX_SITES else "GA",
                "segment": "ASC" if lbl in ASC_SITES else "Clinic",
                "year": yr,
                "net_revenue": round(val, 1),
            })
    df = pd.DataFrame(recs)
    tx25 = df[(df.state == "TX") & (df.year == 2025)].net_revenue.sum()
    assert tx25 > 25000, "TX net revenue tie looks wrong"
    return df


def revenue_drivers(wb) -> pd.DataFrame:
    """Volume vs rate decomposition of YoY net-revenue change (DD1)."""
    k = kpi_annual(wb)[["year", "net_revenue", "procedure_count",
                        "rev_per_procedure", "gross_charges"]].copy()
    k["gross_charge_per_proc"] = k.gross_charges / k.procedure_count
    # build a volume/rate bridge for each YoY step
    bridge = []
    yrs = sorted(k.year.tolist())
    for a, b in zip(yrs[:-1], yrs[1:]):
        pa = k[k.year == a].iloc[0]
        pb = k[k.year == b].iloc[0]
        vol_effect = (pb.procedure_count - pa.procedure_count) * pa.rev_per_procedure / 1000
        rate_effect = (pb.rev_per_procedure - pa.rev_per_procedure) * pb.procedure_count / 1000
        bridge.append({"step": f"{a}->{b}",
                       "start_net_revenue": round(pa.net_revenue, 1),
                       "volume_effect": round(vol_effect, 1),
                       "rate_effect": round(rate_effect, 1),
                       "end_net_revenue": round(pb.net_revenue, 1)})
    return k, pd.DataFrame(bridge)


def ar_monthly(wb) -> pd.DataFrame:
    """Net AR by type from AR sheet (monthly) + DSO using rolling revenue."""
    ar = list(wb["AR"].iter_rows(values_only=True))
    hdr = ar[4]
    cols = list(range(2, 38))                    # monthly Jan23-Dec25
    months = [hdr[j] for j in cols]

    def row_by(label):
        for r in ar:
            if isinstance(r[0], str) and r[0].strip() == label:
                return r
            if len(r) > 1 and isinstance(r[1], str) and r[1].strip() == label:
                return r
        raise KeyError(label)

    std_gross = row_by("Account Receivable - Non-Lien")
    std_res = row_by("Contractual Allowance Reserve - Non-Lien")
    pi_gross = row_by("Accounts Receivable - Lien")
    pi_res = row_by("Contractual Allowance - Lien")
    net_ar = row_by("Gross Accounts Receivable, Net")

    df = pd.DataFrame({
        "month": [m.strftime("%Y-%m") if isinstance(m, dt.datetime) else m for m in months],
        "ar_insurance_gross": [std_gross[j] for j in cols],
        "reserve_insurance": [std_res[j] for j in cols],
        "ar_injury_gross": [pi_gross[j] for j in cols],
        "reserve_injury": [pi_res[j] for j in cols],
        "net_ar": [net_ar[j] for j in cols],
    })
    df["net_ar_insurance"] = df.ar_insurance_gross + df.reserve_insurance
    df["net_ar_injury"] = df.ar_injury_gross + df.reserve_injury
    df["date"] = pd.to_datetime(df.month + "-01")
    return df.sort_values("date").reset_index(drop=True)


def cash_annual(wb) -> pd.DataFrame:
    k = kpi_annual(wb)
    df = k[["year", "gross_charges", "cash_collections",
            "est_future_collections", "net_revenue"]].copy()
    df["cash_conversion_pct"] = df.cash_collections / df.net_revenue * 100
    # year-end net AR + DSO
    ar = ar_monthly(wb)
    ye = {2023: "2023-12", 2024: "2024-12", 2025: "2025-12"}
    df["net_ar_year_end"] = df.year.map(
        lambda y: float(ar[ar.month == ye[y]].net_ar.iloc[0]))
    df["dso_days"] = df.net_ar_year_end / df.net_revenue * 365
    return df


def main():
    wb = load()
    DB.parent.mkdir(parents=True, exist_ok=True)
    if DB.exists():
        DB.unlink()
    con = sqlite3.connect(DB)

    kpi = kpi_annual(wb)
    pnl = pnl_monthly(wb)
    loc = location_annual(wb)
    drivers, bridge = revenue_drivers(wb)
    ar = ar_monthly(wb)
    cash = cash_annual(wb)

    kpi.to_sql("kpi_annual", con, index=False)
    pnl.drop(columns=["date"]).to_sql("pnl_monthly", con, index=False)
    loc.to_sql("location_annual", con, index=False)
    drivers.to_sql("revenue_drivers", con, index=False)
    bridge.to_sql("revenue_bridge", con, index=False)
    ar.drop(columns=["date"]).to_sql("ar_monthly", con, index=False)
    cash.to_sql("cash_annual", con, index=False)

    con.commit()
    con.close()
    print(f"Built {DB}")
    for name, df in [("kpi_annual", kpi), ("pnl_monthly", pnl),
                     ("location_annual", loc), ("revenue_drivers", drivers),
                     ("revenue_bridge", bridge), ("ar_monthly", ar),
                     ("cash_annual", cash)]:
        print(f"  {name:18s} {len(df):>4} rows")


if __name__ == "__main__":
    main()
